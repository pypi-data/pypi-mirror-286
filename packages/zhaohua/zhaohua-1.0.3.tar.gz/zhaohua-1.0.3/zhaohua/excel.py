
#########################################################################################################

# 将一维或二维数组写入到Excel文件 save_to_excel

# 参数:
# - data: 要写入的数据，可以是一维列表或二维列表。
# - file_path: 保存的文件路径（包含路径+文件名+后缀）。如果路径或文件不存在，将自动创建。
# - start_cell: 数据写入的起始位置，可以是以下值之一：
#     - 'begin': 清空现有工作表，从 "A1" 开始写入数据。
#     - 'end': 追加数据到工作表的末尾。
#     - 指定单元格地址，如 "B2": 从指定的单元格开始写入数据。
# - row_or_column: 当数据为一维列表时，'row'表示按行写入，任意其它值表示按列写入。默认为'row'。

#########################################################################################################

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string
import os

def save_to_excel(data, file_path, start_cell='end', row_or_column='row'):
    """
    将一维或二维数组写入到Excel文件。

    参数:
    data (list): 要写入的数据列表。
    file_path (str): Excel文件路径。
    start_cell (str): 数据开始写入的单元格位置。默认为'end'，为追加写入;'begin'为覆盖写入；或指定单元格写入"A5"。
    row_or_column (str): 当数据为一维列表时，'row'表示按行写入，任意其它值表示按列写入。默认为'row'。
    """
    # 检查文件路径是否为空
    if not file_path:
        print("Error: 文件路径不能为空。")
        return

    # 将所有反斜杠替换为斜杠，并确保文件路径完整并且是.xlsx格式
    file_path = os.path.normpath(file_path.replace('\\', '/'))
    base, ext = os.path.splitext(file_path)
    if ext.lower() not in ['.xlsx']:
        file_path = base + '.xlsx'

    # 确保目录存在
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # 检查文件是否存在并加载或创建工作簿
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    # 将start_cell转换为小写以便比较
    start_cell = start_cell.lower()

    # 根据start_cell参数决定数据写入位置
    if start_cell == 'begin':
        # 清除现有工作表数据
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
        row, col = 1, 1  # 从"A1"开始写入
    elif start_cell == 'end':
        # 找到工作表中第一个空白行
        if ws.max_row == 1 and not any(ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)):
            row = 1  # 如果表格为空，从第一行开始写入
        else:
            row = ws.max_row + 1
            while any(ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)):
                row += 1
        col = 1  # 从新的一行的第一列开始写入
    else:
        # 转换start_cell为大写
        start_cell = start_cell.upper()
        # 从特定单元格开始写入
        col_letter = ''.join(filter(str.isalpha, start_cell))
        row = int(''.join(filter(str.isdigit, start_cell)))
        col = column_index_from_string(col_letter)

    # 写入数据
    if isinstance(data[0], list):
        # 如果data的第一个元素是列表，说明data是一个二维列表
        for i, row_data in enumerate(data, start=row):
            # 枚举二维列表中的每一行数据，同时从指定的开始行位置(row)开始计数
            for j, value in enumerate(row_data, start=col):
                # 枚举当前行中的每个值，同时从指定的开始列位置(col)开始计数
                ws.cell(row=i, column=j, value=value)
                # 将值写入到工作表的相应单元格
    else:
        # 如果data的第一个元素不是列表，说明data是一维列表
        if row_or_column == 'row':
            # 如果row_or_column为'row'，按行写入
            for i, value in enumerate(data, start=col):
                # 枚举一维列表中的每个值，同时从指定的开始列位置(col)开始计数
                ws.cell(row=row, column=i, value=value)
                # 将值写入到工作表的相应单元格

        else:
            # 如果row_or_column为其它任意值，按列写入
            for i, value in enumerate(data, start=row):
                # 枚举一维列表中的每个值，同时从指定的开始行位置(row)开始计数
                ws.cell(row=i, column=col, value=value)
                # 将值写入到工作表的相应单元格



    # 保存工作簿
    wb.save(file_path)
    wb.close()