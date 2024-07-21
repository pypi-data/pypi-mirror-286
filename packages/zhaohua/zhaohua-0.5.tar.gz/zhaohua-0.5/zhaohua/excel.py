
#########################################################################################################

# 将一组数组或二维数组写入 save_to_excel

# 参数:
# - data: 要写入的数据，可以是一维列表或二维列表。
# - file_path: 保存的文件路径（包含路径+文件名+后缀）。如果路径或文件不存在，将自动创建。
# - start_cell: 数据写入的起始位置，可以是以下值之一：
#     - 'begin': 清空现有工作表，从 "A1" 开始写入数据。
#     - 'end': 追加数据到工作表的末尾。
#     - 指定单元格地址，如 "B2": 从指定的单元格开始写入数据。

#########################################################################################################

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string
import os

def save_to_excel(data, file_path, start_cell='end'):
    # 检查文件路径是否为空
    if not file_path:
        print("Error: 文件路径不能为空。")
        return

    # 将所有反斜杠替换为斜杠，并确保文件路径完整并且是.xlsx格式
    file_path = os.path.normpath(file_path.replace('\\', '/'))
    base, ext = os.path.splitext(file_path)
    if ext.lower() not in ['.xlsx']:
        file_path = base + '.xlsx'

    # 确保目录存在
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # 检查文件是否存在并加载或创建工作簿
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    # 将start_cell转换为小写以便比较
    start_cell = start_cell.lower()

    # 根据start_cell参数决定数据写入位置
    if start_cell == 'begin':
        # 清除现有工作表数据
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
        row, col = 1, 1  # 从"A1"开始写入
    elif start_cell == 'end':
        # 找到工作表中第一个空白行
        row = ws.max_row + 1 if ws.max_row > 1 else 1
        col = 1  # 从新的一行的第一列开始写入
    else:
        # 转换start_cell为大写
        start_cell = start_cell.upper()
        # 从特定单元格开始写入
        col_letter = ''.join(filter(str.isalpha, start_cell))
        row = int(''.join(filter(str.isdigit, start_cell)))
        col = column_index_from_string(col_letter)

    # 写入数据
    if isinstance(data[0], list):
        for i, row_data in enumerate(data, start=row):
            for j, value in enumerate(row_data, start=col):
                ws.cell(row=i, column=j, value=value)
    else:
        for i, value in enumerate(data, start=row):
            ws.cell(row=i, column=col, value=value)

    # 保存工作簿
    wb.save(file_path)
    wb.close()
