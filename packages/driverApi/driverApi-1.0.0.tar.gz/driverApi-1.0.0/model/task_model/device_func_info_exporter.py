# author: haoliqing
# date: 2024/1/8 17:18
# desc:
from model.function.function_config import FunctionConfig
from model.global_param_manager import GlobalParamManager
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill


class DeviceFuncInfoExporter(object):
    manager: GlobalParamManager = GlobalParamManager()

    def export(self):
        wb = openpyxl.Workbook()
        list_ws = wb.create_sheet("功能列表")
        list_ws.cell(row=1, column=1, value="序号")
        list_ws.cell(row=1, column=2, value="功能名称")
        list_ws.cell(row=1, column=3, value="功能描述")
        func_cfgs = self.manager.device_func_cfgs
        func_counter = 0
        for func in func_cfgs.values():
            func_counter += 1
            func_name = func.function_name
            func_desc = func.function_desc
            # row 和 column 的值都是从1开始
            list_ws.cell(row=func_counter + 1, column=1, value=func_counter)
            list_ws.cell(row=func_counter + 1, column=2, value=func_name)
            list_ws.cell(row=func_counter + 1, column=3, value=func_desc)
            func_ws = wb.create_sheet(func_name)
            func_ws.cell(row=1, column=1, value="功能名称")
            func_ws.cell(row=1, column=2, value=func_name)
            func_ws.cell(row=2, column=1, value="功能描述")
            func_ws.cell(row=2, column=2, value=func_desc)
            func_ws.cell(row=3, column=1, value="参数序号")
            func_ws.cell(row=3, column=2, value="参数名称")
            func_ws.cell(row=3, column=3, value="参数描述")
            params: list[dict] = func.get_param_cfg()
            param_counter = 0
            for param in params:
                param_counter += 1
                paramName = param[FunctionConfig.PARAM_NODE_NAME]
                paramDesc = param[FunctionConfig.PARAM_NODE_DESC]
                func_ws.cell(row=param_counter + 3, column=1, value=param_counter)
                func_ws.cell(row=param_counter + 3, column=2, value=paramName)
                func_ws.cell(row=param_counter + 3, column=3, value=paramDesc)

        default_sheet = wb["Sheet"]
        if default_sheet:
            wb.remove(default_sheet)
        wb.save("D://外设功能接口定义.xlsx")

    def export2(self):
        wb = openpyxl.Workbook()
        list_ws_id = "功能列表"
        list_ws = wb.create_sheet(list_ws_id)
        list_ws.append(["序号", "功能名称", "功能描述"])
        title_row = list_ws.row_dimensions[1]
        title_row.height = 20
        cell_ids = ['A1', 'B1', 'C1']
        for cell_id in cell_ids:
            list_ws[cell_id].font = Font(bold=True, size=12)
            list_ws[cell_id].fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FF000000')
        list_ws.column_dimensions['B'].width = 30
        list_ws.column_dimensions['C'].width = 100
        func_cfgs = self.manager.device_func_cfgs
        func_counter = 0
        for func in func_cfgs.values():
            func_counter += 1
            func_name = func.function_name
            func_desc = func.function_desc
            list_ws.append([str(func_counter), func_name, func_desc])
            # 加一个超链接
            name_cell_id = 'B' + str(func_counter + 1)
            self.add_inner_hyper_link(list_ws, name_cell_id, func_name)

            func_ws = wb.create_sheet(func_name)
            func_ws.column_dimensions['A'].width = 10
            func_ws.column_dimensions['B'].width = 30
            func_ws.column_dimensions['C'].width = 10
            func_ws.column_dimensions['D'].width = 15
            func_ws.column_dimensions['E'].width = 100
            func_ws.append(["<< 返回"])
            back_cell_id = 'A1'
            self.add_inner_hyper_link(func_ws, back_cell_id, list_ws_id)
            func_ws.append(["功能名称", func_name])
            func_ws.append(["功能描述", func_desc])
            func_ws.append(["参数序号", "参数名称", "是否必输", "是否返回值", "参数描述"])
            cell_ids = ['A4', 'B4', 'C4', 'D4', 'E4']
            for cell_id in cell_ids:
                func_ws[cell_id].font = Font(bold=True, size=12)
                func_ws[cell_id].fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FF000000')
            params: list[dict] = func.get_param_cfg()
            param_counter = 0
            for param in params:
                param_counter += 1
                paramName = param[FunctionConfig.PARAM_NODE_NAME]
                paramDesc = param.get(FunctionConfig.PARAM_NODE_DESC, None)
                paramRequire = param.get(FunctionConfig.PARAM_NODE_CHECK_REQUIRE, None)
                paramReturn = param.get(FunctionConfig.PARAM_NODE_RETURN_VALUE, None)
                func_ws.append([str(param_counter), paramName, paramRequire, paramReturn, paramDesc])

        default_sheet = wb["Sheet"]
        if default_sheet:
            wb.remove(default_sheet)
        wb.save("D://外设功能接口定义.xlsx")

    def add_inner_hyper_link(self, source_ws, source_cell_id, target_ws_id):
        """
        添加文档内的超链接
        :param source_ws:  要添加超链接的work_sheet
        :param source_cell_id:  要添加超链接的单元格ID
        :param target_ws_id: 链接目标work_sheet_id
        :return:
        """
        cell = source_ws[source_cell_id]
        cell.hyperlink = '#' + target_ws_id + '!' + source_cell_id
        cell.font = Font(underline='single', color='0000FF')


if __name__ == '__main__':
    DeviceFuncInfoExporter().export2()
