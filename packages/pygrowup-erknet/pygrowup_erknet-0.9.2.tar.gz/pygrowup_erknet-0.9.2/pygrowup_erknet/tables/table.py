import requests
import csv
from openpyxl import load_workbook
from pathlib import Path
import io
from enum import Enum
import math
from typing import Dict, Union

from .lms import LMS
from .sex import Sex
from .math import z_to_percentile, percentile_to_z, calculate_value_by_percentile
from ..exceptions import DataNotFoundError

cache_dir = Path(__file__).resolve().parent / 'cache'

class TableType(str, Enum):
    BodyMassIndexForAge = 'bmifa'
    HeadCircumferenceForAge = 'hcfa'
    LengthForAge = 'lfa'
    HeightForAge = 'hfa'
    WeightForAge = 'wfa'
    WeightForHeigt = 'wfh'
    WeightForLength = 'wfl'

TYPES_ALIASES = {
    TableType.LengthForAge : TableType.HeightForAge,
    TableType.WeightForLength : TableType.WeightForHeigt
}

AGE_BASED_TYPES = [
    TableType.BodyMassIndexForAge,
    TableType.HeadCircumferenceForAge,
    TableType.LengthForAge,
    TableType.HeightForAge,
    TableType.WeightForAge,
]

LENGTH_BASED_TYPES = [
    TableType.WeightForHeigt,
    TableType.WeightForLength,
]

class GrowthTable(object):
    _table: Union[Dict[Sex, Dict[float, LMS]], None] = None
    _min_range: Union[float, None] = None
    _max_range: Union[float, None] = None

    def __init__(self) -> None:
        self.get_table()
    
    def get_lms(self, index_value: Union[float, int], sex: Union[Sex, None] = None) -> LMS:
        lms = None
        table_by_sex = self.get_table()[sex] if self.get_col_index_sex() is not None else self.get_table()[None]
        if (index_value >= self.get_min_range()) and (index_value <= self.get_max_range()):
            closest_index = min(table_by_sex.keys(), key=lambda k: abs(k - index_value))
            lms = table_by_sex.get(closest_index, None)
        if lms is None:
            raise DataNotFoundError(index_value)
        return lms
    
    def calculate_zscore(self, index_value: Union[float, int], measurement: Union[float, int], sex: Union[Sex, None] = None) -> float:
        ###
        # calculate z-score
        # https://www.cdc.gov/growthcharts/extended-bmi-data-files.htm
        #
        #           [y/M(t)]^L(t) - 1
        #   Zind =  -----------------
        #               S(t)L(t)
        ###
        lms = self.get_lms(index_value, sex)
        zscore = 0.0
        if (lms.l == 0):
            zscore = math.log((measurement / lms.m), math.e) / lms.s
        else:
            zscore = ((measurement / lms.m) ** lms.l - 1) / (lms.l * lms.s)
        if (self.get_col_index_sigma() is None) or (z_to_percentile(zscore) <= 0.95):
            return zscore
        else:
            percentile = 0.9 + 0.1 * z_to_percentile((measurement - calculate_value_by_percentile(0.9, lms)) / lms.sigma)
            return percentile_to_z(percentile)
    
    def calculate_sd(self, lms: LMS, position: Union[int, float]) -> float:
        return lms.m * ((1 + lms.l * lms.s * position) ** (1 / lms.l))
    
    def get_table_type(self) -> TableType:
        raise NotImplementedError()
    
    def get_table_sex(self) -> Union[Sex, None]:
        return None

    def get_file_name(self) -> str:
        raise NotImplementedError()
    
    def get_csv_url(self) -> str:
        raise NotImplementedError()

    def get_xslx_url(self) -> str:
        raise NotImplementedError()
    
    def get_col_index_sex(self) -> Union[int, None]:
        return 0
    
    def get_col_sex_male_value(self) -> Union[int, None]:
        return 1
    
    def get_col_index_measurement(self) -> int:
        return 1
    
    def get_col_index_l(self) -> int:
        return 2
    
    def get_col_index_m(self) -> int:
        return 3
    
    def get_col_index_s(self) -> int:
        return 4
    
    def get_col_index_sigma(self) -> Union[int, None]:
        return None
    
    def get_index_factor(self) -> Union[float, None]:
        return None
    
    def save_csv(self, content: bytes) -> None:
        file_path = cache_dir / f'{self.get_file_name()}.csv'
        file_path.parent.mkdir(parents=True, exist_ok=True)
        with file_path.open('wb') as out_file:
            out_file.write(content)
        
    def download_csv(self) -> None:
        response = requests.get(url=self.get_csv_url(), verify=True)
        if response.status_code == 200:
            self.save_csv(response.content)
    
    def download_xlsx(self) -> None:
        response = requests.get(url=self.get_xslx_url(), verify=True)
        if response.status_code == 200:
            wb = load_workbook(io.BytesIO(response.content))
            ws = wb.worksheets[0]
            csv_stream = io.StringIO()
            csv_writer = csv.writer(csv_stream)
            for row in ws.iter_rows(values_only=True):
                csv_writer.writerow(row)
            self.save_csv(csv_stream.getvalue().encode('utf-8'))
    
    def download_file(self) -> None:
        try:
            self.download_csv()
        except NotImplementedError:
            self.download_xlsx()
    
    def get_table(self) -> Dict[Sex, Dict[float, LMS]]:
        if self._table is not None:
            return self._table
        
        file_path = cache_dir / f'{self.get_file_name()}.csv'
        
        if not file_path.exists():
            self.download_file()
        
        table = {}
        with file_path.open('r') as csvfile:
            reader = csv.reader(csvfile)
            next(reader)  # Skip header
            for row in reader:
                sex = None
                failed = False
                if self.get_col_index_sex() is not None:
                    try:
                        sex = Sex.Male if int(row[self.get_col_index_sex()]) == self.get_col_sex_male_value() else Sex.Female
                    except ValueError:
                        failed = True
                if not failed:
                    if sex not in table:
                        table[sex] = {}
                    index = 0.0
                    try:
                        index = float(row[self.get_col_index_measurement()])
                    except ValueError:
                        failed = True
                    if not failed:
                        if self.get_index_factor() is not None:
                            index = index * self.get_index_factor()
                        sigma = None if self.get_col_index_sigma() is None else float(row[self.get_col_index_sigma()])
                        table[sex][index] = LMS(float(row[self.get_col_index_l()]), float(row[self.get_col_index_m()]), float(row[self.get_col_index_s()]), sigma)
        self._table = table
        return self._table
    
    def get_min_range(self) -> float:
        if self._min_range is None:
            first_item = next(iter(self.get_table().values()), None)
            self._min_range = min(first_item.keys())
        return self._min_range
    
    def get_max_range(self) -> float:
        if self._max_range is None:
            first_item = next(iter(self.get_table().values()), None)
            self._max_range = max(first_item.keys())
        return self._max_range
